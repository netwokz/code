import openpyxl

# path = "C:\\Users\\deanejst\\Documents\\code\\python\\tac-segments-test.xlsx"
path = "C:\\Users\\netwokz\\Documents\\CODE\\code\\python\\tac-segments-test.xlsx"

ALLOWED_ASSETS = ["AmbaFlex", "Best Flex", "Control Cabinets", "Dematic SC3 Sorter", "Dematic SL2 Sorter", "Directed Palletization", "Electrical Conveyors", "Intralox", "Mechanical Conveyors", "RWC4 (Fanuc-R2000, Robotic work cells)", "SLAM", "Shop Tools", "SmartPac", "Tote Stacker", "Transnorm"
                  ]

TAC_101 = "101"
TAC_201 = "201"
TAC_301 = "301"

headers_to_remove = ["Site", "ESD"]

fhn = ["qmoychri", "mperezf", "deanejst", "feticher", "breadj", "ahuertaj"]

headers = []
asset = []
segment = []
name = []
login = []
completed = []

wb_object = openpyxl.load_workbook(path)
sheet_obj = wb_object.active
row = sheet_obj.max_row
column = sheet_obj.max_column

print("\nValue of first column")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    # headers.append(cell_obj.value)
    # print(cell_obj.value)

# Cycle through the headers and remove unwanted columns.
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    if cell_obj.value in headers_to_remove:
        sheet_obj.delete_cols(i)

# Cycle through the rows and remove unwanted rows.
for i in range(2, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    if cell_obj.value not in enumerate(reversed(ALLOWED_ASSETS)):
        print(f"Deleting Row {i} {cell_obj.value}")
        sheet_obj.delete_rows(i, 1)

wb_object.save(
    "C:\\Users\\netwokz\\Documents\\CODE\\code\\python\\modified.xlsx")
# path = "C:\\Users\\deanejst\\Documents\\code\\python\\modified.xlsx"

# for i in range(1, row + 1):
#     cell_obj = sheet_obj.cell(row = i, column = 2)
#     if cell_obj.value in ALLOWED_ASSETS:
#         print(cell_obj.value)
