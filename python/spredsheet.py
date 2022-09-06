import openpyxl
import math

path = "C:\\Users\\deanejst\\Documents\\code\\python\\sheet.xlsx"

hours = []

wb_object = openpyxl.load_workbook(path)

sheet_obj = wb_object.active

row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows: ", row)

for i in range(2, row + 1):
    cell_obj = sheet_obj.cell(row = i, column=11)
    # print(cell_obj.value)
    hours.append(float(cell_obj.value))

print("Total Hours: ", math.fsum(hours))